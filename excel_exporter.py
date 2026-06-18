from email import header
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import PieChart, Reference
def export_to_excel(test_cases):

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"output/test_cases_{timestamp}.xlsx"
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Test Cases"

    summary_sheet = workbook.create_sheet(title="Summary")

    headers = ["Test ID", "Scenario", "Expected Result", "Priority"]
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")  

    # define a simple thin border for data cells
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    
    for test_case in test_cases:
        sheet.append(test_case.to_list())

        row = sheet.max_row

        priority_cell = sheet[f"D{row}"]

        if test_case.priority == "High":
            priority_cell.fill = PatternFill(
            "solid",
            start_color="FF9999"
        )

        elif test_case.priority == "Medium":
            priority_cell.fill = PatternFill(
            "solid",
            start_color="FFE599"
        )

        elif test_case.priority == "Low":
            priority_cell.fill = PatternFill(
            "solid",
            start_color="B6D7A8"
        )
    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=4):
        for cell in row:
            cell.border = border
    for column in sheet.columns:

        max_length = 0

        column_letter = column[0].column_letter

        for cell in column:

            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = max_length + 2

        sheet.column_dimensions[
            column_letter
        ].width = adjusted_width

    sheet.auto_filter.ref = sheet.dimensions

    high_count = 0
    medium_count = 0
    low_count = 0

    for test_case in test_cases:
        if test_case.priority == "High":
            high_count += 1
        elif test_case.priority == "Medium":
            medium_count += 1
        elif test_case.priority == "Low":
            low_count += 1

    summary_sheet.append(["Priority", "Count"])
    summary_sheet.append(["High", high_count])
    summary_sheet.append(["Medium", medium_count])
    summary_sheet.append(["Low", low_count])

    summary_sheet.append(["Total", len(test_cases)])
    summary_sheet.append(["High Priority Percentage", f"{(high_count / len(test_cases)) * 100:.2f}%"])
    summary_sheet.append(["Medium Priority Percentage", f"{(medium_count / len(test_cases)) * 100:.2f}%"])
    summary_sheet.append(["Low Priority Percentage", f"{(low_count / len(test_cases)) * 100:.2f}%"])

    pie = PieChart()

    labels = Reference(summary_sheet, min_col=1, min_row=3, max_row=5)
    data = Reference(summary_sheet, min_col=2, min_row=3, max_row=5)

    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Test Case Priority Distribution"

    summary_sheet.add_chart(pie, "D2")

    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)

        cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

        summary_sheet["A3"].fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        summary_sheet["A4"].fill = PatternFill(start_color="FFE599", end_color="FFE599", fill_type="solid")
        summary_sheet["A5"].fill = PatternFill(start_color="B6D7A8", end_color="B6D7A8", fill_type="solid")
    for column in summary_sheet.columns:

        max_length = 0

        column_letter = column[0].column_letter

        for cell in column:

            if cell.value:

                max_length = max(max_length, len(str(cell.value)))

        summary_sheet.column_dimensions[column_letter].width = max_length + 2

    workbook.save(filename)
    return filename
